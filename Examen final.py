
from openpyxl import Workbook

wb = Workbook()

# Seleccionar la hoja activa
hoja = wb.active
hoja.title = "Agenda"

# Agregar datos a celdas específicas
hoja['A1'] = "Dominio"
hoja['B1'] = "Reino"
hoja['C1'] = "División"
hoja['D1'] = "Clase"
hoja['E1'] = "Orden"
hoja['F1'] = "Familia"
hoja['G1'] = "Género"
hoja['H1'] = "Especie"
hoja['I1'] = "Nombre científico"
hoja['J1'] = "uuid"
hoja['K1'] = "etag"



# Guardar el archivo
wb.save("plantas.xlsx")


from openpyxl import load_workbook

# Cargar el archivo Excel
wb = load_workbook("ejemplo.xlsx")

# Seleccionar una hoja
hoja = wb["Agenda"]

# Leer datos específicos de una celda
nombre = hoja["A2"].value  # Lee el valor de la celda A2
edad = hoja["B2"].value    # Lee el valor de la celda B2
ciudad = hoja["C2"].value  # Lee el valor de la celda C2

print(f"Nombre: {nombre}, Edad: {edad}, Ciudad: {ciudad}")

# Leer todos los datos de la hoja
for fila in hoja.iter_rows(min_row=1, max_col=3, values_only=True):
    print(fila)
hoja.merge_cells("A1:D1")
import matplotlib 